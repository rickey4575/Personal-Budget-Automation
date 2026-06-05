#!/usr/bin/env python3
"""
budget_tool.py
Local, privacy-friendly budget automation:
- Read one or many statement files (XLSX/CSV)
- Normalize columns (supports your statement columns like Type, Started Date, Completed Date, Description, Amount, Fee)
- Auto-categorize with keyword rules
- Append to a master ledger (CSV)
- Update one master budget workbook using Budget_Template.xlsx

Usage:
  python budget_tool.py --input statements
  python budget_tool.py --input path/to/statement.xlsx --workbook Budget_Master.xlsx

Requirements:
  pip install pandas openpyxl python-dateutil
  pip install google-genai  # optional, only for AI categorization
"""

import argparse
import csv
import json
import os
import re
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.views import Selection

HERE = Path(__file__).resolve().parent

def load_config(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def load_rules(path: Path):
    rules = []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            kw = (row.get("keyword") or "").strip()
            cat = (row.get("category") or "").strip()
            if kw and cat:
                rules.append((kw.lower(), cat))
    return rules

def append_category_rules(path: Path, new_rules: dict[str, str]) -> int:
    if not new_rules:
        return 0

    existing = set()
    if path.exists():
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                keyword = (row.get("keyword") or "").strip().lower()
                if keyword:
                    existing.add(keyword)

    rows = [
        {"keyword": keyword, "category": category}
        for keyword, category in sorted(new_rules.items(), key=lambda item: item[0].lower())
        if keyword.strip().lower() not in existing
    ]
    if not rows:
        return 0

    needs_header = not path.exists() or path.stat().st_size == 0
    with open(path, "a", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["keyword", "category"])
        if needs_header:
            writer.writeheader()
        writer.writerows(rows)

    return len(rows)

def list_files(input_path: Path):
    if input_path.is_file():
        return [input_path]
    files = []
    for ext in ("*.xlsx", "*.xlsm", "*.csv"):
        files.extend(sorted(input_path.glob(ext)))
    return files

def _lower_map(df: pd.DataFrame):
    return {str(c).strip().lower(): c for c in df.columns}

def _find_col(df: pd.DataFrame, candidates):
    lm = _lower_map(df)
    # exact match
    for cand in candidates:
        cand = cand.lower().strip()
        if cand in lm:
            return lm[cand]
    # contains / fuzzy
    for cand in candidates:
        cand = cand.lower().strip()
        for lc, orig in lm.items():
            if cand in lc:
                return orig
    return None

def _to_number(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce").fillna(0)

def normalize_columns(df: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    """
    Output columns: Date, Description, Amount
    Optionally uses:
      - Separate debit/credit columns: amount = credit - debit
      - Fee column: net amount = amount - fee (if fee_handling == 'subtract')
      - Type column: if amounts appear all positive, infer sign for debits/credits
    """
    cm = cfg.get("column_mapping", {})
    # Prefer completed date if present, else started date, else generic date
    date_col = _find_col(df, cm.get("date", ["date"]))
    desc_col = _find_col(df, cm.get("description", ["description"]))
    fee_col  = _find_col(df, cm.get("fee", ["fee"])) if cm.get("fee") else None
    amount_style = cfg.get("amount_style", "signed")

    if not (date_col and desc_col):
        raise ValueError(f"Could not detect required columns. Found: date={date_col}, description={desc_col}")

    out = df[[date_col, desc_col]].copy()
    out.columns = ["Date", "Description"]

    if amount_style == "debit_credit":
        dc = cfg.get("debit_credit_columns", {})
        debit_col = _find_col(df, dc.get("debit", ["debit"]))
        credit_col = _find_col(df, dc.get("credit", ["credit"]))

        if not (debit_col or credit_col):
            raise ValueError(f"Could not detect debit or credit columns. Found: debit={debit_col}, credit={credit_col}")

        debit = _to_number(df[debit_col]) if debit_col else 0
        credit = _to_number(df[credit_col]) if credit_col else 0
        out["Amount"] = credit - debit
    else:
        amt_col = _find_col(df, cm.get("amount", ["amount"]))
        if not amt_col:
            raise ValueError(f"Could not detect amount column. Found: amount={amt_col}")
        out["Amount"] = pd.to_numeric(df[amt_col], errors="coerce")

    # Fee handling (optional)
    if fee_col:
        fee = _to_number(df[fee_col])
        amt = pd.to_numeric(out["Amount"], errors="coerce")
        if cfg.get("fee_handling", "ignore") == "subtract":
            out["Amount"] = amt - fee
        else:
            out["Amount"] = amt

    return out

def parse_dates(s: pd.Series, hint: str = "") -> pd.Series:
    if hint:
        return pd.to_datetime(s, errors="coerce", format=hint)
    return pd.to_datetime(s, errors="coerce")

def categorize(desc: str, rules) -> str:
    d = (desc or "").lower()
    for kw, cat in rules:
        if kw in d:
            return cat
    if any(x in d for x in ["payroll", "salary", "wages"]):
        return "Salary"
    if any(x in d for x in ["transfer", "zelle", "venmo", "cashapp", "cash app"]):
        return "Transfers"
    if any(x in d for x in ["fee", "overdraft"]):
        return "Fees"
    return "Uncategorized"

def month_str(dt: pd.Timestamp) -> str:
    if pd.isna(dt):
        return ""
    return dt.strftime("%Y-%m")

def detect_latest_month(df: pd.DataFrame) -> str:
    months = df["Month"].dropna().astype(str)
    months = [m for m in months if re.match(r"^\d{4}-\d{2}$", m)]
    return sorted(months)[-1] if months else ""

def clean_transactions(df: pd.DataFrame, rules=None) -> pd.DataFrame:
    tx = df.copy()
    tx["Date"] = parse_dates(tx["Date"])
    tx = tx.dropna(subset=["Date"])
    tx["Description"] = tx["Description"].astype(str)
    tx["Amount"] = pd.to_numeric(tx["Amount"], errors="coerce")
    tx = tx.dropna(subset=["Amount"])
    tx["Month"] = tx["Date"].apply(month_str)
    if rules is not None:
        tx["Category"] = tx["Description"].apply(lambda d: categorize(d, rules))
    elif "Category" not in tx.columns:
        tx["Category"] = "Uncategorized"
    return tx[["Date", "Description", "Amount", "Category", "Month"]]

def merge_ledger(old: pd.DataFrame, new: pd.DataFrame, rules) -> pd.DataFrame:
    merged = pd.concat([old, new], ignore_index=True)
    merged = clean_transactions(merged, rules)
    merged = merged.sort_values(["Date", "Description", "Amount"])
    return merged.drop_duplicates(subset=["Date", "Description", "Amount"], keep="last")

def maybe_infer_sign_from_type(original_df: pd.DataFrame, tx: pd.DataFrame, cfg: dict) -> pd.DataFrame:
    """
    If Amount column is all non-negative (common in some exports),
    infer debit as negative using Type column values.
    """
    type_candidates = cfg.get("type_column", [])
    if not type_candidates:
        return tx

    type_col = _find_col(original_df, type_candidates)
    if not type_col:
        return tx

    amt = pd.to_numeric(tx["Amount"], errors="coerce")
    if amt.dropna().empty:
        return tx

    # Only apply if amounts look unsigned (>=0)
    if (amt.dropna() >= 0).all():
        t = original_df[type_col].astype(str).str.lower().str.strip()
        debit_vals = [v.lower() for v in cfg.get("debit_type_values", ["debit"])]
        credit_vals = [v.lower() for v in cfg.get("credit_type_values", ["credit"])]

        is_debit = t.apply(lambda x: any(v in x for v in debit_vals))
        is_credit = t.apply(lambda x: any(v in x for v in credit_vals))

        signed = amt.copy()
        signed[is_debit] = -signed[is_debit]
        # credits stay positive; unknown types stay as-is
        tx["Amount"] = signed

    return tx

def load_transactions(input_path: Path, cfg: dict, rules) -> pd.DataFrame:
    files = list_files(input_path)
    if not files:
        raise SystemExit(f"No .xlsx/.xlsm/.csv files found in {input_path}")

    all_tx = []
    for fp in files:
        if fp.suffix.lower() == ".csv":
            df = pd.read_csv(fp)
        else:
            df = pd.read_excel(fp)

        tx = normalize_columns(df, cfg)

        tx["Date"] = parse_dates(tx["Date"], cfg.get("date_format_hint",""))
        tx = tx.dropna(subset=["Date"])

        tx["Description"] = tx["Description"].astype(str)
        tx["Amount"] = pd.to_numeric(tx["Amount"], errors="coerce")
        tx = tx.dropna(subset=["Amount"])

        tx = maybe_infer_sign_from_type(df, tx, cfg)

        tx["Category"] = tx["Description"].apply(lambda d: categorize(d, rules))
        tx["Month"] = tx["Date"].apply(month_str)

        all_tx.append(tx)

    return pd.concat(all_tx, ignore_index=True)

def available_categories(rules, ledger: pd.DataFrame | None = None):
    categories = {cat for _, cat in rules}
    if ledger is not None and "Category" in ledger.columns:
        categories.update(str(c) for c in ledger["Category"].dropna().unique())
    categories.discard("Uncategorized")
    return sorted(categories)

def uncategorized_groups(tx: pd.DataFrame, limit: int) -> list[dict]:
    uncategorized = tx[tx["Category"] == "Uncategorized"].copy()
    if uncategorized.empty:
        return []

    uncategorized["Amount"] = pd.to_numeric(uncategorized["Amount"], errors="coerce").fillna(0)
    grouped = (
        uncategorized
        .groupby("Description", as_index=False)
        .agg(
            count=("Description", "size"),
            total_amount=("Amount", "sum"),
            total_spend=("Amount", lambda s: float(abs(s[s < 0].sum()))),
            first_seen=("Date", "min"),
            last_seen=("Date", "max"),
        )
        .sort_values(["count", "total_spend"], ascending=False)
        .head(limit)
    )

    groups = []
    for _, row in grouped.iterrows():
        groups.append({
            "merchant": row["Description"],
            "count": int(row["count"]),
            "total_amount": round(float(row["total_amount"]), 2),
            "total_spend": round(float(row["total_spend"]), 2),
            "first_seen": row["first_seen"].strftime("%Y-%m-%d") if hasattr(row["first_seen"], "strftime") else str(row["first_seen"]),
            "last_seen": row["last_seen"].strftime("%Y-%m-%d") if hasattr(row["last_seen"], "strftime") else str(row["last_seen"]),
        })
    return groups

def build_gemini_prompt(groups: list[dict], categories: list[str]) -> str:
    return json.dumps({
        "task": "Suggest budget categories for uncategorized bank transaction merchant descriptions.",
        "instructions": [
            "Use an existing category whenever it reasonably fits.",
            "Suggest a new category only when no existing category is suitable.",
            "Use concise category names.",
            "Return valid JSON only, with no markdown.",
            "Do not include personal judgments or sensitive commentary.",
        ],
        "existing_categories": categories,
        "required_response_shape": {
            "suggestions": [
                {
                    "merchant": "exact merchant string from input",
                    "suggested_category": "category name",
                    "confidence": 0.0,
                    "create_new_category": False,
                    "reason": "short explanation",
                }
            ]
        },
        "uncategorized_merchants": groups,
    }, indent=2)

def _extract_json_object(text: str) -> dict:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    start = text.find("{")
    end = text.rfind("}")
    if start == -1 or end == -1 or end < start:
        raise ValueError("Gemini did not return a JSON object")
    return json.loads(text[start:end + 1])

def request_gemini_suggestions(groups: list[dict], categories: list[str], model: str) -> list[dict]:
    if not os.environ.get("GEMINI_API_KEY"):
        raise SystemExit("GEMINI_API_KEY is not set. Add it to your environment before using --suggest-categories.")

    try:
        from google import genai
    except ImportError as exc:
        raise SystemExit("google-genai is not installed. Run: pip install google-genai") from exc

    client = genai.Client()
    response = client.models.generate_content(
        model=model,
        contents=build_gemini_prompt(groups, categories),
    )
    parsed = _extract_json_object(response.text or "")
    suggestions = parsed.get("suggestions", [])
    if not isinstance(suggestions, list):
        raise ValueError("Gemini response did not include a suggestions list")
    return suggestions

def _boolish(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "y"}

def build_ai_decision_rows(groups: list[dict], suggestions: list[dict], categories: list[str]):
    by_merchant = {str(s.get("merchant", "")).strip(): s for s in suggestions if isinstance(s, dict)}
    existing_categories = {c.strip().lower(): c for c in categories}
    rows = []
    apply_map = {}
    for group in groups:
        merchant = group["merchant"]
        suggestion = by_merchant.get(merchant, {})
        suggested_category = str(suggestion.get("suggested_category", "")).strip()
        category_key = suggested_category.lower()
        creates_new = _boolish(suggestion.get("create_new_category", False))
        matched_existing = category_key in existing_categories
        applied = bool(suggested_category and matched_existing and not creates_new)

        if applied:
            suggested_category = existing_categories[category_key]
            apply_map[merchant] = suggested_category

        rows.append({
            "merchant": merchant,
            "count": group["count"],
            "total_spend": group["total_spend"],
            "total_amount": group["total_amount"],
            "first_seen": group["first_seen"],
            "last_seen": group["last_seen"],
            "suggested_category": suggested_category,
            "confidence": suggestion.get("confidence", ""),
            "create_new_category": creates_new,
            "matched_existing_category": matched_existing,
            "applied": applied,
            "reason": suggestion.get("reason", ""),
            "needs_approval": not applied,
            "approved": "",
        })

    return rows, apply_map

def write_category_suggestions(path: Path, groups: list[dict], suggestions: list[dict], categories: list[str]):
    rows, _ = build_ai_decision_rows(groups, suggestions, categories)

    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_csv(path, index=False)

def apply_ai_categories(tx: pd.DataFrame, apply_map: dict[str, str]) -> pd.DataFrame:
    if not apply_map:
        return tx

    updated = tx.copy()
    mask = updated["Category"].eq("Uncategorized") & updated["Description"].isin(apply_map)
    updated.loc[mask, "Category"] = updated.loc[mask, "Description"].map(apply_map)
    return updated

def run_ai_category_flow(ledger: pd.DataFrame, rules, limit: int, model: str, log_path: Path):
    groups = uncategorized_groups(ledger, limit)
    if not groups:
        return ledger, {}, 0

    categories = available_categories(rules, ledger)
    suggestions = request_gemini_suggestions(groups, categories, model)
    rows, apply_map = build_ai_decision_rows(groups, suggestions, categories)

    log_path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_csv(log_path, index=False)

    updated = apply_ai_categories(ledger, apply_map)
    needs_approval = sum(1 for row in rows if row["needs_approval"])
    return updated, apply_map, needs_approval

def _sheet(wb, name: str):
    return wb[name] if name in wb.sheetnames else wb.create_sheet(name)

def _clear_sheet(ws):
    if ws.max_row:
        ws.delete_rows(1, ws.max_row)

def _replace_rows_after_header(ws, headers):
    _clear_sheet(ws)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

def _period_actual_formula(category_cell: str) -> str:
    return (
        f'=IF($B$1="All",'
        f'ABS(SUMIFS(Transactions!$C:$C,Transactions!$D:$D,{category_cell},Transactions!$C:$C,"<0")),'
        f'ABS(SUMIFS(Transactions!$C:$C,Transactions!$D:$D,{category_cell},Transactions!$E:$E,$B$1,Transactions!$C:$C,"<0")))'
    )

def _period_metric_formula(kind: str) -> str:
    if kind == "income":
        return '=IF($B$1="All",SUMIFS(Transactions!$C:$C,Transactions!$C:$C,">0",Transactions!$D:$D,"<>Transfers"),SUMIFS(Transactions!$C:$C,Transactions!$E:$E,$B$1,Transactions!$C:$C,">0",Transactions!$D:$D,"<>Transfers"))'
    if kind == "spending":
        return '=IF($B$1="All",ABS(SUMIFS(Transactions!$C:$C,Transactions!$C:$C,"<0",Transactions!$D:$D,"<>Transfers")),ABS(SUMIFS(Transactions!$C:$C,Transactions!$E:$E,$B$1,Transactions!$C:$C,"<0",Transactions!$D:$D,"<>Transfers")))'
    if kind == "uncategorized_spend":
        return '=IF($B$1="All",ABS(SUMIFS(Transactions!$C:$C,Transactions!$C:$C,"<0",Transactions!$D:$D,"Uncategorized")),ABS(SUMIFS(Transactions!$C:$C,Transactions!$E:$E,$B$1,Transactions!$C:$C,"<0",Transactions!$D:$D,"Uncategorized")))'
    if kind == "uncategorized_count":
        return '=IF($B$1="All",COUNTIFS(Transactions!$C:$C,"<0",Transactions!$D:$D,"Uncategorized"),COUNTIFS(Transactions!$E:$E,$B$1,Transactions!$C:$C,"<0",Transactions!$D:$D,"Uncategorized"))'
    if kind == "transfer_in":
        return '=IF($B$1="All",SUMIFS(Transactions!$C:$C,Transactions!$C:$C,">0",Transactions!$D:$D,"Transfers"),SUMIFS(Transactions!$C:$C,Transactions!$E:$E,$B$1,Transactions!$C:$C,">0",Transactions!$D:$D,"Transfers"))'
    if kind == "transfer_out":
        return '=IF($B$1="All",ABS(SUMIFS(Transactions!$C:$C,Transactions!$C:$C,"<0",Transactions!$D:$D,"Transfers")),ABS(SUMIFS(Transactions!$C:$C,Transactions!$E:$E,$B$1,Transactions!$C:$C,"<0",Transactions!$D:$D,"Transfers")))'
    raise ValueError(f"Unknown metric formula: {kind}")

def _style_dashboard_section(ws, row: int, fill_color: str):
    for cell in ws[row][0:5]:
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center")

def _style_dashboard_table_header(ws, row: int):
    for cell in ws[row][0:5]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True, color="000000")
        cell.alignment = Alignment(vertical="center")

def _reset_sheet_view(ws, active_cell: str = "A1"):
    ws.freeze_panes = None
    ws.sheet_view.pane = None
    ws.sheet_view.topLeftCell = None
    ws.sheet_view.selection = [Selection(activeCell=active_cell, sqref=active_cell)]

def _configure_period_dropdown(wb, months, selected_period: str):
    ws = _sheet(wb, "Periods")
    _replace_rows_after_header(ws, ["Period"])
    for period in ["All", *months]:
        ws.append([period])
    ws.sheet_state = "hidden"

    dashboard = wb["Dashboard"]
    dashboard["A1"] = "Selected Period:"
    dashboard["B1"] = selected_period if selected_period in ["All", *months] else (months[-1] if months else "All")
    dashboard["D1"] = "Use the dropdown in B1 to choose a month or All"

    for dv in list(dashboard.data_validations.dataValidation):
        if "B1" in str(dv.sqref):
            dashboard.data_validations.dataValidation.remove(dv)

    end_row = max(2, len(months) + 2)
    dv = DataValidation(type="list", formula1=f"=Periods!$A$2:$A${end_row}", allow_blank=False)
    dashboard.add_data_validation(dv)
    dv.add(dashboard["B1"])

def _ensure_budget_categories(wb, tx: pd.DataFrame):
    ws = wb["Budget"]
    if ws.max_row == 0:
        ws.append(["Category", "Monthly Budget"])

    existing = {
        str(ws.cell(row, 1).value).strip().lower()
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 1).value
    }
    skip = {"salary", "other income", "transfers", "uncategorized"}
    categories = sorted(
        str(category)
        for category in tx["Category"].dropna().unique()
        if str(category).strip().lower() not in skip
    )

    for category in categories:
        if category.strip().lower() not in existing:
            ws.append([category, 0])
            existing.add(category.strip().lower())

def _update_dashboard(wb, tx: pd.DataFrame):
    ws = wb["Dashboard"]
    _reset_sheet_view(ws)

    ws["A3"] = "Key Metrics"
    ws["A4"] = "Total Income"
    ws["B4"] = _period_metric_formula("income")
    ws["A5"] = "Total Spending"
    ws["B5"] = _period_metric_formula("spending")
    ws["A6"] = "Net Cashflow"
    ws["B6"] = "=B4-B5"
    ws["A7"] = "Savings Rate"
    ws["B7"] = "=IFERROR(B6/B4,0)"

    if ws.max_row >= 9:
        ws.delete_rows(9, ws.max_row - 8)

    ws["A9"] = "Review"
    ws["A10"] = "Uncategorized Spend"
    ws["B10"] = _period_metric_formula("uncategorized_spend")
    ws["A11"] = "Uncategorized Count"
    ws["B11"] = _period_metric_formula("uncategorized_count")

    ws["A13"] = "Transfers"
    ws["A14"] = "Transfer In"
    ws["B14"] = _period_metric_formula("transfer_in")
    ws["A15"] = "Transfer Out"
    ws["B15"] = _period_metric_formula("transfer_out")
    ws["A16"] = "Net Transfers"
    ws["B16"] = "=B14-B15"

    budget_start = 20
    ws["A18"] = "Budget vs Actual (Spending)"
    ws["A19"] = "Category"
    ws["B19"] = "Budget"
    ws["C19"] = "Actual"
    ws["D19"] = "Difference"
    ws["E19"] = "Status"

    budget_ws = wb["Budget"]
    categories = []
    for row in range(2, budget_ws.max_row + 1):
        category = budget_ws.cell(row, 1).value
        if category:
            categories.append(str(category))

    if not categories:
        categories = sorted(c for c in tx["Category"].dropna().unique() if c not in ["Salary", "Transfers"])

    for offset, category in enumerate(categories, start=budget_start):
        ws.cell(offset, 1).value = category
        ws.cell(offset, 2).value = f'=IFERROR(VLOOKUP(A{offset},Budget!$A:$B,2,FALSE),0)'
        ws.cell(offset, 3).value = _period_actual_formula(f"A{offset}")
        ws.cell(offset, 4).value = f"=B{offset}-C{offset}"
        ws.cell(offset, 5).value = f'=IF(D{offset}<0,"Over Budget","OK")'

    for row in ws.iter_rows(min_row=1, max_row=19, max_col=5):
        for cell in row:
            if cell.row in [1, 3, 9, 13, 18, 19]:
                cell.font = Font(bold=True)

    _style_dashboard_section(ws, 3, "1F4E78")
    _style_dashboard_section(ws, 9, "70AD47")
    _style_dashboard_section(ws, 13, "8064A2")
    _style_dashboard_section(ws, 18, "5B9BD5")
    _style_dashboard_table_header(ws, 19)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

def _update_uncategorized(wb, tx: pd.DataFrame):
    ws = _sheet(wb, "Uncategorized")
    _replace_rows_after_header(ws, ["Date", "Description", "Amount", "Category", "Month"])

    uncategorized = tx[tx["Category"] == "Uncategorized"].sort_values(["Month", "Date", "Description"])
    for _, row in uncategorized.iterrows():
        dt = row["Date"]
        ws.append([
            dt.date() if hasattr(dt, "date") else dt,
            row["Description"],
            float(row["Amount"]),
            row["Category"],
            row["Month"],
        ])

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12

def update_master_workbook(template_path: Path, workbook_path: Path, tx: pd.DataFrame, selected_period: str):
    if workbook_path.exists():
        wb = load_workbook(workbook_path)
    else:
        wb = load_workbook(template_path)

    wsT = wb["Transactions"]
    _replace_rows_after_header(wsT, ["Date", "Description", "Amount", "Category", "Month"])

    for _, row in tx.iterrows():
        dt = row["Date"]
        wsT.append([dt.date() if hasattr(dt, "date") else dt, row["Description"], float(row["Amount"]), row["Category"], row["Month"]])

    months = sorted(m for m in tx["Month"].dropna().astype(str).unique() if re.match(r"^\d{4}-\d{2}$", m))
    _ensure_budget_categories(wb, tx)
    _configure_period_dropdown(wb, months, selected_period)
    _update_dashboard(wb, tx)
    _update_uncategorized(wb, tx)

    workbook_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_path)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="Path to a statement file or a folder of statement files")
    ap.add_argument("--output", help="Legacy option; reports now update --workbook instead of writing monthly output files")
    ap.add_argument("--workbook", default=str(HERE / "Budget_Master.xlsx"), help="Path to the master budget workbook")
    ap.add_argument("--template", default=str(HERE / "Budget_Template.xlsx"), help="Path to Budget_Template.xlsx")
    ap.add_argument("--config", default=str(HERE / "config.json"), help="Path to config.json")
    ap.add_argument("--rules", default=str(HERE / "category_rules.csv"), help="Path to category_rules.csv")
    ap.add_argument("--ledger", default=str(HERE / "master_ledger.csv"), help="Path to master ledger CSV (will be created/updated)")
    ap.add_argument("--suggest-categories", action="store_true", help="Ask Gemini for category suggestions for uncategorized merchant groups")
    ap.add_argument("--use-ai-categories", action="store_true", help="Automatically apply Gemini suggestions that match existing categories")
    ap.add_argument("--suggestions-output", default=str(HERE / "ai_category_suggestions.csv"), help="Where to write review-only AI category suggestions")
    ap.add_argument("--ai-log", default=str(HERE / "ai_category_decisions.csv"), help="Where to write the AI categorization audit log")
    ap.add_argument("--suggestion-limit", type=int, default=500, help="Maximum uncategorized merchant groups to send to Gemini")
    ap.add_argument("--gemini-model", default="gemini-3.1-flash-lite", help="Gemini model to use for category suggestions")
    args = ap.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    workbook_path = Path(args.workbook).expanduser().resolve()
    template_path = Path(args.template).expanduser().resolve()
    cfg = load_config(Path(args.config).expanduser().resolve())
    rules = load_rules(Path(args.rules).expanduser().resolve())
    ledger_path = Path(args.ledger).expanduser().resolve()

    tx_all = load_transactions(input_path, cfg, rules)

    if ledger_path.exists():
        old = pd.read_csv(ledger_path)
        ledger = merge_ledger(old, tx_all, rules)
    else:
        ledger = clean_transactions(tx_all, rules)

    if args.suggest_categories:
        groups = uncategorized_groups(ledger, args.suggestion_limit)
        if not groups:
            print("No uncategorized merchant groups found.")
            return
        categories = available_categories(rules, ledger)
        suggestions = request_gemini_suggestions(groups, categories, args.gemini_model)
        suggestions_path = Path(args.suggestions_output).expanduser().resolve()
        write_category_suggestions(suggestions_path, groups, suggestions, categories)
        print(f"Wrote category suggestions: {suggestions_path}")
        print("Review the suggestions before adding any rows to category_rules.csv.")
        return

    if args.use_ai_categories:
        ai_log_path = Path(args.ai_log).expanduser().resolve()
        ledger, apply_map, needs_approval_count = run_ai_category_flow(
            ledger,
            rules,
            args.suggestion_limit,
            args.gemini_model,
            ai_log_path,
        )
        rules_path = Path(args.rules).expanduser().resolve()
        added_rule_count = append_category_rules(rules_path, apply_map)
        if added_rule_count:
            rules = load_rules(rules_path)
            ledger = clean_transactions(ledger, rules)
        print(f"Wrote AI categorization log: {ai_log_path}")
        print(f"Applied existing-category AI decisions: {len(apply_map)}")
        print(f"Added category rules: {added_rule_count}")
        if needs_approval_count:
            print(f"AI decisions needing approval: {needs_approval_count}")

    ledger.to_csv(ledger_path, index=False)

    latest = detect_latest_month(tx_all) or detect_latest_month(ledger) or datetime.today().strftime("%Y-%m")

    update_master_workbook(template_path, workbook_path, ledger, latest)

    print(f"Updated workbook: {workbook_path}")
    print(f"Updated ledger: {ledger_path}")
    print("Tip: open the workbook and check 'Uncategorized' to improve your rules.")

if __name__ == "__main__":
    main()
